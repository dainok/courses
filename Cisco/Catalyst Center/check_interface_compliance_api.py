#!/usr/bin/env python3
import re
import requests
import yaml
import openpyxl
from ttp import ttp

total_interfaces = 0
compliant_interfaces = 0


def compare_interface(template, interface, ignore_value=None):
    """Compare an interface against a profile."""
    if not ignore_value:
        # Set default
        ignore_value = []

    for key, value in template.items():
        if key not in interface:
            # template config does not exist in interface
            return False
        if key in ignore_value and interface[key] is not None:
            # tempate config value exists, value is ignored
            continue
        if interface[key] is False and value is not False:
            # template requires False
            return False
        if interface[key] is True and value is not True:
            # template requires True
            return False
        if interface[key] != value:
            # tempate config value differs from interface
            return False
        # print(f"{key}|{value}|{interface[key]}")
    return True


summary_headers = [
    "type",
    "name",
    "percentage",
    "compliant",
    "total",
]
xlsx_headers = [
    "device",
    "interface",
    "profile",
    "is_compliant",
    "disabled",
    "access_vlan",
    "voice_vlan",
    "swport_mode",
    "swport_negotiate",
    "tracking_policy",
    "load_interval",
    "stp_portfast",
    "stp_guardroot",
    "description",
]
interfaces = []
profiles = {}
summary = []


# Load config file
with open("config.yml") as fh:
    config = yaml.safe_load(fh)


# Connect to DNAC
url = f"{config['dnac_url']}/dna/system/api/v1/auth/token"
response = requests.post(
    url,
    auth=(config["dnac_username"], config["dnac_password"]),
    verify=config["verify_cert"],
    timeout=30,
)
headers = {"Accept": "application/json", "X-Auth-Token": response.json()["Token"]}


# Get devices
url = f"{config['dnac_url']}/dna/intent/api/v1/network-device"
device_list = requests.get(
    url, headers=headers, verify=config["verify_cert"], timeout=30
).json()["response"]
device_details = {}
for item in device_list:
    if item["family"] != "Switches and Hubs":
        # Skip non switch devices
        continue

    id = item["id"]
    hostname = item["hostname"].upper()
    os_type = item["softwareType"]
    os_version = item["softwareVersion"]
    config_url = f"{config['dnac_url']}/dna/intent/api/v1/network-device/{id}/config"
    running_config = requests.get(
        config_url, headers=headers, verify=config["verify_cert"], timeout=30
    ).json()["response"]
    device_details[hostname] = {
        "id": id,
        "os_type": os_type,
        "os_version": os_version,
        "running_config": running_config,
    }


# Parsing config
with open("interface.ttp", "r") as fh:
    ttp_template = fh.read()
for hostname, details in device_details.items():
    parser = ttp(data=device_details[hostname]["running_config"], template=ttp_template)
    parser.parse()
    device_details[hostname]["interface_config"] = parser.result()[0][0]["interfaces"]


# Getting templates
for profile in config["profiles"]:
    # Get template config (use template against interface profile)
    template_parser = ttp(data=profile["interface_template"], template=ttp_template)
    template_parser.parse()
    template_config = template_parser.result()[0][0]["interfaces"]

    # Save template
    profiles[profile["name"]] = template_config


# Checking interface against profiles
for hostname, details in device_details.items():
    # For each device
    for interface in details["interface_config"]:
        # For each interface
        full_interface_name = f"{hostname}:{interface['interface']}"

        # Find template
        profile_name = None
        is_compliant = False
        for profile in config["profiles"]:
            if re.match(f"^{profile['interface_regex']}$", full_interface_name):
                # Profile match, update counter
                total_interfaces = total_interfaces + 1

                # Test interface compliance
                is_compliant = compare_interface(
                    profiles[profile["name"]],
                    interface,
                    ignore_value=config["ignore_value"],
                )
                if is_compliant:
                    # Interface is compliant, update counter
                    compliant_interfaces = compliant_interfaces + 1
                    print(f"{full_interface_name} is compliant")
                else:
                    print(f"{full_interface_name} is not compliant")

                # Profile checked, proceed with next interface
                break

        # Add attributes
        interface["device"] = hostname
        interface["profile"] = profile["name"]
        interface["is_compliant"] = is_compliant

        # Transform interface to a list of ordered attributes
        new_interface = []
        for key in xlsx_headers:
            new_interface.append(interface.get(key))
        interfaces.append(new_interface)


# Calcolating the summary
overall_total_interfaces = 0
overall_compliant_interfaces = 0
device_total_interfaces = 0
device_compliant_interfaces = 0
current_device = None
for interface in interfaces:
    device = interface[0]
    if not current_device:
        # Initial set
        current_device = device
    has_profile = True if interface[2] else False
    is_compliant = interface[3]

    if current_device and device != current_device:
        # Add to summary and reset counters
        summary.append(
            [
                "device",
                current_device,
                (
                    int(device_compliant_interfaces / device_total_interfaces * 100)
                    if device_total_interfaces > 0
                    else "NA"
                ),
                device_compliant_interfaces,
                device_total_interfaces,
            ]
        )
        current_device = device
        device_total_interfaces = 0
        device_compliant_interfaces = 0

    if has_profile:
        overall_total_interfaces = overall_total_interfaces + 1
        device_total_interfaces = device_total_interfaces + 1

        if is_compliant:
            overall_compliant_interfaces = overall_compliant_interfaces + 1
            device_compliant_interfaces = device_compliant_interfaces + 1
else:
    if current_device:
        # Append last items
        summary.append(
            [
                "device",
                current_device,
                (
                    int(device_compliant_interfaces / device_total_interfaces * 100)
                    if device_total_interfaces > 0
                    else "NA"
                ),
                device_compliant_interfaces,
                device_total_interfaces,
            ]
        )
        summary.append(
            [
                "overall",
                "",
                (
                    int(overall_compliant_interfaces / overall_total_interfaces * 100)
                    if overall_total_interfaces > 0
                    else "NA"
                ),
                overall_compliant_interfaces,
                overall_total_interfaces,
            ]
        )


# Create XLSXxlsx
wb = openpyxl.Workbook()
summary_ws = wb.active
summary_ws.title = "Summary"
interface_ws = wb.create_sheet("Interfaces")
profile_ws = wb.create_sheet("Profiles")

# Add headers
interface_ws.append(xlsx_headers)
profile_ws.append(xlsx_headers)
summary_ws.append(summary_headers)

# Add data
for line in summary:
    summary_ws.append(line)
for interface in interfaces:
    interface_ws.append(interface)
for name, profile in profiles.items():
    profile["profile"] = name
    profile_ws.append(profile.get(key) for key in xlsx_headers)

# Add filters
interface_ws.auto_filter.ref = interface_ws.dimensions
profile_ws.auto_filter.ref = profile_ws.dimensions
summary_ws.auto_filter.ref = profile_ws.dimensions

# Auto size
for ws in [summary_ws, interface_ws, profile_ws]:
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

# Save XLSX
wb.save("export.xlsx")
